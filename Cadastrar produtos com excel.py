import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment

produtos = []

# Função para cadastrar um novo produto
def cadastrar_produto():
    print("Cadastro de Produto:")
    print("Escolha o tipo de produto:")

    # Opções de tipos de produto
    opcoes = {
        '1': 'SSD',
        '2': 'Memória RAM',
        '3': 'Placa Mãe',
        '4': 'Processador',
        '5': 'Placa de Vídeo',
        '6': 'Gabinete',
        '7': 'Fonte',
        '8': 'Cooler',
        '9': 'Fans'
    }

    # Mostra as opções de tipo de produto
    for key, value in opcoes.items():
        print(f"{key} - {value}")

    # Obtém a escolha do usuário
    escolha = input("Escolha o número correspondente ao tipo de produto: ")

    # Verifica se a escolha é válida
    if escolha in opcoes:
        tipo_produto = opcoes[escolha]
        marca = input("Marca do Produto: ")

        # Pergunta se é kit ou unidade e trata conforme a resposta
        preco_por_kit = input("O preço é por kit? (Sim/Não): ").lower() == 'sim'

        if preco_por_kit:
            quantidade = int(input("Quantidade no kit: "))
            valor_total = float(input("Valor total do kit: "))
            valor_unitario = valor_total / quantidade
        else:
            valor_unitario = float(input("Valor unitário: "))
            valor_total = valor_unitario  # Se for por unidade, o valor total é o mesmo que o valor unitário

        dataCompra = input("Data de compra (01/01/1999): ")

        if escolha == '1':  # SSD
            capacidade = input("Capacidade (ex: 500GB): ")
            tecnologia = input("Tecnologia (ex: SATA, NVMe): ")
            velocidade_leitura = input("Velocidade de Leitura (ex: 550MB/s): ")

            observacao = input("Observação: ")

            produtos.append({
                'tipo': tipo_produto,
                'marca': marca,
                'valor_unitario': valor_unitario,
                'valor_total': valor_total,
                'quantidade': quantidade if preco_por_kit else 1,
                'dataCompra': dataCompra,
                'capacidade': capacidade,
                'tecnologia': tecnologia,
                'velocidade_leitura': velocidade_leitura,
                'observacao': observacao
            })

        elif escolha == '2':  # Memória RAM
            capacidade = input("Capacidade (ex: 8GB): ")
            frequencia = input("Frequência (ex: 3200MHz): ")
            cor = input("Cor: ")
            rgb = input("É RGB? (Sim/Não): ")

            observacao = input("Observação: ")

            produtos.append({
                'tipo': tipo_produto,
                'marca': marca,
                'valor_unitario': valor_unitario,
                'valor_total': valor_total,
                'quantidade': quantidade if preco_por_kit else 1,
                'dataCompra': dataCompra,
                'capacidade': capacidade,
                'frequencia': frequencia,
                'cor': cor,
                'rgb': rgb.lower() == 'sim',
                'observacao': observacao
            })

        else:  # Outros tipos de produtos
            observacao = input("Observação: ")

            produtos.append({
                'tipo': tipo_produto,
                'marca': marca,
                'valor_unitario': valor_unitario,
                'valor_total': valor_total,
                'quantidade': quantidade if preco_por_kit else 1,
                'dataCompra': dataCompra,
                'observacao': observacao
            })

        print("Produto cadastrado com sucesso\n")
    else:
        print("Opção inválida. Tente novamente.")

# Função para listar todos os produtos cadastrados
def listar_produtos():
    print("Listagem de Produtos")
    if not produtos:
        print("Nenhum produto cadastrado.")
    else:
        produtos_ordenados = sorted(produtos, key=lambda x: x['tipo'])
        for produto in produtos_ordenados:
            if produto['quantidade'] > 1:  # Se for por kit, exibe a quantidade e o valor unitário
                print(f"Tipo: {produto['tipo']}, Marca: {produto['marca']}, Quantidade: {produto['quantidade']}, Valor Unitário: R$ {produto['valor_unitario']:.2f}, Valor Total: R$ {produto['valor_total']:.2f}, Data de Compra: {produto['dataCompra']}")
            else:  # Se for por unidade, exibe apenas o valor unitário
                print(f"Tipo: {produto['tipo']}, Marca: {produto['marca']}, Valor Unitário: R$ {produto['valor_unitario']:.2f}, Data de Compra: {produto['dataCompra']}")

            if produto['tipo'] == 'SSD':
                print(f"Capacidade: {produto['capacidade']}, Tecnologia: {produto['tecnologia']}, Velocidade de Leitura: {produto['velocidade_leitura']}, Observação: {produto['observacao']}")
            elif produto['tipo'] == 'Memória RAM':
                rgb_status = 'Sim' if produto['rgb'] else 'Não'
                print(f"Capacidade: {produto['capacidade']}, Frequência: {produto['frequencia']}, Cor: {produto['cor']}, RGB: {rgb_status}, Observação: {produto['observacao']}")
            else:
                print(f"Observação: {produto['observacao']}")

# Função para exportar os produtos para uma planilha Excel
def exportar_para_excel():
    if not produtos:
        print("Não há produtos cadastrados para exportar.")
        return

    try:
        # Solicita o nome do arquivo Excel
        nome_arquivo = input("Digite o nome do arquivo Excel onde deseja exportar (exemplo.xlsx): ")
        # Carrega o arquivo Excel existente
        workbook = load_workbook(nome_arquivo)
    except FileNotFoundError:
        print(f"O arquivo '{nome_arquivo}' não foi encontrado.")
        return
    except Exception as e:
        print(f"Erro ao abrir o arquivo '{nome_arquivo}': {str(e)}")
        return

    # Verifica se a planilha 'Produtos' já existe no arquivo, senão cria uma nova
    if 'Produtos' in workbook.sheetnames:
        sheet = workbook['Produtos']
    else:
        sheet = workbook.create_sheet(title='Produtos')

        # Cabeçalho da planilha
        sheet.append(["Tipo", "Marca", "Preço Unitário", "Quantidade", "Preço Total", "Data de Compra", "Capacidade", "Tecnologia", "Velocidade de Leitura", "Frequência", "Cor", "RGB", "Observação"])

    # Adiciona os produtos à planilha
    for produto in produtos:
        if produto['quantidade'] > 1:  # Se for por kit
            sheet.append([
                produto['tipo'],
                produto['marca'],
                produto['valor_unitario'],
                produto['quantidade'],
                produto['valor_total'],
                produto['dataCompra'],
                produto.get('capacidade', ''),
                produto.get('tecnologia', ''),
                produto.get('velocidade_leitura', ''),
                produto.get('frequencia', ''),
                produto.get('cor', ''),
                'Sim' if produto.get('rgb', False) else 'Não',
                produto['observacao']
            ])
        else:  # Se for por unidade
            sheet.append([
                produto['tipo'],
                produto['marca'],
                produto['valor_unitario'],
                '',
                '',
                produto['dataCompra'],
                produto.get('capacidade', ''),
                produto.get('tecnologia', ''),
                produto.get('velocidade_leitura', ''),
                produto.get('frequencia', ''),
                produto.get('cor', ''),
                'Sim' if produto.get('rgb', False) else 'Não',
                produto['observacao']
            ])

    # Ajusta a largura das colunas
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column].width = adjusted_width

    # Salva o arquivo Excel
    workbook.save(nome_arquivo)
    print(f"Dados exportados com sucesso para o arquivo '{nome_arquivo}'.")

# Função principal do programa
def main():
    while True:
        print("\nEscolha uma opção:")
        print("1 - Cadastrar novo produto")
        print("2 - Listar produtos")
        print("3 - Exportar para Excel")
        print("4 - Sair")
        opcao = input("Opção: ")

        if opcao == '1':
            cadastrar_produto()
        elif opcao == '2':
            listar_produtos()
        elif opcao == '3':
            exportar_para_excel()
        elif opcao == '4':
            print("Saindo...")
            break
        else:
            print("Opção inválida. Tente novamente.")

if __name__ == "__main__":
    main()
